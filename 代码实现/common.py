from __future__ import annotations

import csv
import heapq
import math
import random
from collections import deque
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import matplotlib

# 使用非交互式后端，避免在无 GUI 环境下保存图片时报错。
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PROJECT_DIR = Path(__file__).resolve().parent
DATA_SOURCE_DIR = ROOT / "A题-2026年长春理工大学大学生数学建模竞赛题目" / "A题-2026年长春理工大学大学生数学建模竞赛题目"
ATTACHMENT_1 = DATA_SOURCE_DIR / "附件一.xlsx"
ATTACHMENT_2 = DATA_SOURCE_DIR / "附件二.xlsx"
RESULTS_DIR = PROJECT_DIR / "results"
FIGURES_DIR = PROJECT_DIR / "figures"
DATA_DIR = PROJECT_DIR / "data"
ACCEPTANCE_DIR = PROJECT_DIR / "acceptance"

HEIGHT = 400
WIDTH = 234
DEFAULT_CELL_SIZE_KM = 0.25


@dataclass
class Uav:
    uav_id: str
    start_x: float
    start_y: float
    max_flight_time: float
    max_speed: float
    sensor_range: float
    comm_range: float
    fuel_consumption: float
    status: int


@dataclass
class Target:
    target_id: int
    x: float
    y: float
    weight: float
    cell_count: int
    local_risk: float
    service_time: float = 1.0


@dataclass
class RouteResult:
    uav_id: str
    sequence: List[int]
    time: float
    reward: float
    expected_reward: float
    risk_cost: float
    survival: float
    energy: float
    feasible: bool
    path_points: List[Tuple[float, float]]


def ensure_dirs() -> None:
    for path in (RESULTS_DIR, FIGURES_DIR, DATA_DIR, ACCEPTANCE_DIR):
        path.mkdir(parents=True, exist_ok=True)


def setup_plot_style() -> None:
    plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "DejaVu Sans", "Arial"]
    plt.rcParams["axes.unicode_minus"] = False
    plt.rcParams["font.size"] = 10
    plt.rcParams["axes.linewidth"] = 1.2
    plt.rcParams["xtick.major.width"] = 1.2
    plt.rcParams["ytick.major.width"] = 1.2
    plt.rcParams["lines.linewidth"] = 2


def read_sheet_matrix(path: Path, sheet_name: str, max_rows: int, max_cols: int) -> np.ndarray:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    data = np.zeros((max_rows, max_cols), dtype=float)
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=max_cols, values_only=True)
    ):
        for col_idx, value in enumerate(row):
            data[row_idx, col_idx] = 0.0 if value is None else float(value)
    wb.close()
    return data


def load_uavs(path: Path = ATTACHMENT_2) -> List[Uav]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    header = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {name: i for i, name in enumerate(header)}
    uavs: List[Uav] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[index["uav_id"]] is None or row[index["start_x"]] is None:
            continue
        if not str(row[index["uav_id"]]).startswith("UAV-"):
            continue
        try:
            uavs.append(
                Uav(
                    uav_id=str(row[index["uav_id"]]),
                    start_x=float(row[index["start_x"]]),
                    start_y=float(row[index["start_y"]]),
                    max_flight_time=float(row[index["max_flight_time"]]),
                    max_speed=float(row[index["max_speed"]]),
                    sensor_range=float(row[index["sensor_range"]]),
                    comm_range=float(row[index["comm_range"]]),
                    fuel_consumption=float(row[index["fuel_consumption"]]),
                    status=int(row[index["status"]]),
                )
            )
        except (TypeError, ValueError):
            continue
    wb.close()
    return uavs


def load_base_data() -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    base_risk = read_sheet_matrix(ATTACHMENT_1, "各网格风险基础概率", HEIGHT, WIDTH)
    air_defense = read_sheet_matrix(ATTACHMENT_1, "七处防空火力部署区域", HEIGHT, WIDTH)
    value = read_sheet_matrix(ATTACHMENT_1, "网格区域价值", HEIGHT, WIDTH)
    return base_risk, air_defense, value


def find_fire_points(air_defense: np.ndarray) -> List[Tuple[int, int]]:
    ys, xs = np.where(air_defense > 0)
    return [(int(x), int(y)) for y, x in zip(ys, xs)]


def build_risk_field(
    base_risk: np.ndarray,
    fire_points: Sequence[Tuple[int, int]],
    amplitude: float = 0.65,
    sigma: float = 18.0,
) -> np.ndarray:
    yy, xx = np.indices(base_risk.shape)
    safe_product = 1.0 - np.clip(base_risk, 0.0, 0.999)
    for fx, fy in fire_points:
        add = amplitude * np.exp(-((xx - fx) ** 2 + (yy - fy) ** 2) / (2.0 * sigma**2))
        safe_product *= 1.0 - np.clip(add, 0.0, 0.95)
    return np.clip(1.0 - safe_product, 0.0, 0.999)


def extract_targets(value: np.ndarray, risk: np.ndarray, threshold: float = 0.1) -> List[Target]:
    mask = value > threshold
    visited = np.zeros(mask.shape, dtype=bool)
    targets: List[Target] = []
    next_id = 1
    height, width = mask.shape
    for y in range(height):
        for x in range(width):
            if not mask[y, x] or visited[y, x]:
                continue
            queue: deque[Tuple[int, int]] = deque([(x, y)])
            visited[y, x] = True
            cells: List[Tuple[int, int]] = []
            while queue:
                cx, cy = queue.popleft()
                cells.append((cx, cy))
                for nx, ny in ((cx + 1, cy), (cx - 1, cy), (cx, cy + 1), (cx, cy - 1)):
                    if 0 <= nx < width and 0 <= ny < height and mask[ny, nx] and not visited[ny, nx]:
                        visited[ny, nx] = True
                        queue.append((nx, ny))
            weights = np.array([value[cy, cx] for cx, cy in cells], dtype=float)
            total_weight = float(weights.sum())
            xs = np.array([cx for cx, _ in cells], dtype=float)
            ys = np.array([cy for _, cy in cells], dtype=float)
            if total_weight > 0:
                center_x = float((xs * weights).sum() / total_weight)
                center_y = float((ys * weights).sum() / total_weight)
            else:
                center_x = float(xs.mean())
                center_y = float(ys.mean())
            local_risk = float(np.mean([risk[cy, cx] for cx, cy in cells]))
            targets.append(
                Target(
                    target_id=next_id,
                    x=center_x,
                    y=center_y,
                    weight=total_weight,
                    cell_count=len(cells),
                    local_risk=local_risk,
                    service_time=1.0,
                )
            )
            next_id += 1
    targets.sort(key=lambda t: t.weight, reverse=True)
    return targets


def grid_distance(a: Tuple[float, float], b: Tuple[float, float]) -> float:
    return math.hypot(a[0] - b[0], a[1] - b[1])


def point_distance(a: Tuple[float, float], b: Tuple[float, float], cell_size_km: float = DEFAULT_CELL_SIZE_KM) -> float:
    return grid_distance(a, b) * cell_size_km


def sample_line(a: Tuple[float, float], b: Tuple[float, float]) -> List[Tuple[int, int]]:
    dist = max(1, int(math.ceil(grid_distance(a, b) * 2)))
    points: List[Tuple[int, int]] = []
    for idx in range(dist + 1):
        t = idx / dist
        x = int(round(a[0] * (1 - t) + b[0] * t))
        y = int(round(a[1] * (1 - t) + b[1] * t))
        x = min(max(x, 0), WIDTH - 1)
        y = min(max(y, 0), HEIGHT - 1)
        if not points or points[-1] != (x, y):
            points.append((x, y))
    return points


def segment_stats(
    risk: np.ndarray,
    start: Tuple[float, float],
    end: Tuple[float, float],
    speed: float,
    alpha_risk: float = 2.0,
    cell_size_km: float = DEFAULT_CELL_SIZE_KM,
) -> Dict[str, float]:
    if grid_distance(start, end) < 1e-9:
        return {"distance": 0.0, "time": 0.0, "risk_cost": 0.0, "survival": 1.0, "cost": 0.0}
    line = sample_line(start, end)
    risks = np.array([risk[y, x] for x, y in line], dtype=float)
    dist = point_distance(start, end, cell_size_km=cell_size_km)
    risk_cost = float(-np.log(1.0 - np.clip(risks, 0.0, 0.999)).sum())
    return {
        "distance": dist,
        "time": dist / speed,
        "risk_cost": risk_cost,
        "survival": math.exp(-risk_cost),
        "cost": dist / speed + alpha_risk * risk_cost,
    }


def evaluate_sequence(
    sequence: Sequence[Target],
    risk: np.ndarray,
    start: Tuple[float, float],
    speed: float,
    fuel_consumption: float,
    max_time: float,
    service_time: float = 1.0,
    cell_size_km: float = DEFAULT_CELL_SIZE_KM,
) -> RouteResult:
    points: List[Tuple[float, float]] = [start] + [(t.x, t.y) for t in sequence] + [start]
    total_time = 0.0
    risk_cost = 0.0
    total_reward = 0.0
    expected_reward = 0.0
    survival_to_now = 1.0
    path_points = [start]
    for idx in range(len(points) - 1):
        stats = segment_stats(risk, points[idx], points[idx + 1], speed, cell_size_km=cell_size_km)
        total_time += stats["time"]
        risk_cost += stats["risk_cost"]
        survival_to_now *= stats["survival"]
        path_points.append(points[idx + 1])
        if idx < len(sequence):
            target = sequence[idx]
            total_time += service_time
            total_reward += target.weight
            expected_reward += target.weight * survival_to_now
    energy = fuel_consumption * sum(point_distance(points[i], points[i + 1], cell_size_km=cell_size_km) for i in range(len(points) - 1))
    return RouteResult(
        uav_id="",
        sequence=[t.target_id for t in sequence],
        time=total_time,
        reward=total_reward,
        expected_reward=expected_reward,
        risk_cost=risk_cost,
        survival=math.exp(-risk_cost),
        energy=energy,
        feasible=total_time <= max_time,
        path_points=path_points,
    )


def greedy_insert_route(
    candidates: Sequence[Target],
    risk: np.ndarray,
    start: Tuple[float, float],
    speed: float,
    max_time: float,
    fuel_consumption: float,
    max_targets: int = 20,
    cell_size_km: float = DEFAULT_CELL_SIZE_KM,
) -> List[Target]:
    route: List[Target] = []
    remaining = list(candidates)
    while remaining and len(route) < max_targets:
        best_score = -1.0
        best_target: Optional[Target] = None
        best_pos = 0
        for target in remaining:
            for pos in range(len(route) + 1):
                trial = route[:pos] + [target] + route[pos:]
                result = evaluate_sequence(trial, risk, start, speed, fuel_consumption, max_time, cell_size_km=cell_size_km)
                if not result.feasible:
                    continue
                extra_time = max(
                    result.time - evaluate_sequence(route, risk, start, speed, fuel_consumption, max_time, cell_size_km=cell_size_km).time,
                    1e-6,
                )
                score = target.weight * result.survival / extra_time
                if score > best_score:
                    best_score = score
                    best_target = target
                    best_pos = pos
        if best_target is None:
            break
        route.insert(best_pos, best_target)
        remaining = [t for t in remaining if t.target_id != best_target.target_id]
    return two_opt_route(route, risk, start, speed, fuel_consumption, max_time, cell_size_km=cell_size_km)


def two_opt_route(
    route: Sequence[Target],
    risk: np.ndarray,
    start: Tuple[float, float],
    speed: float,
    fuel_consumption: float,
    max_time: float,
    max_iter: int = 80,
    cell_size_km: float = DEFAULT_CELL_SIZE_KM,
) -> List[Target]:
    best = list(route)
    best_value = evaluate_sequence(best, risk, start, speed, fuel_consumption, max_time, cell_size_km=cell_size_km).expected_reward
    improved = True
    rounds = 0
    while improved and rounds < max_iter:
        improved = False
        rounds += 1
        for i in range(len(best) - 1):
            for j in range(i + 2, len(best) + 1):
                trial = best[:i] + list(reversed(best[i:j])) + best[j:]
                result = evaluate_sequence(trial, risk, start, speed, fuel_consumption, max_time, cell_size_km=cell_size_km)
                if result.feasible and result.expected_reward > best_value + 1e-9:
                    best = trial
                    best_value = result.expected_reward
                    improved = True
                    break
            if improved:
                break
    return best


def astar_path(
    risk: np.ndarray,
    start: Tuple[float, float],
    goal: Tuple[float, float],
    p_safe: float = 0.85,
    alpha_risk: float = 4.0,
) -> List[Tuple[int, int]]:
    sx, sy = int(round(start[0])), int(round(start[1]))
    gx, gy = int(round(goal[0])), int(round(goal[1]))
    sx, sy = min(max(sx, 0), WIDTH - 1), min(max(sy, 0), HEIGHT - 1)
    gx, gy = min(max(gx, 0), WIDTH - 1), min(max(gy, 0), HEIGHT - 1)
    neighbors = [
        (1, 0, 1.0),
        (-1, 0, 1.0),
        (0, 1, 1.0),
        (0, -1, 1.0),
        (1, 1, math.sqrt(2)),
        (1, -1, math.sqrt(2)),
        (-1, 1, math.sqrt(2)),
        (-1, -1, math.sqrt(2)),
    ]
    start_node = (sx, sy)
    goal_node = (gx, gy)
    open_heap: List[Tuple[float, Tuple[int, int]]] = [(0.0, start_node)]
    came_from: Dict[Tuple[int, int], Tuple[int, int]] = {}
    g_score: Dict[Tuple[int, int], float] = {start_node: 0.0}
    closed = set()
    while open_heap:
        _, current = heapq.heappop(open_heap)
        if current in closed:
            continue
        if current == goal_node:
            path = [current]
            while current in came_from:
                current = came_from[current]
                path.append(current)
            return list(reversed(path))
        closed.add(current)
        cx, cy = current
        for dx, dy, dist in neighbors:
            nx, ny = cx + dx, cy + dy
            if nx < 0 or nx >= WIDTH or ny < 0 or ny >= HEIGHT:
                continue
            if risk[ny, nx] > p_safe and (nx, ny) != goal_node:
                continue
            step_cost = dist + alpha_risk * float(risk[ny, nx])
            tentative = g_score[current] + step_cost
            node = (nx, ny)
            if tentative < g_score.get(node, float("inf")):
                came_from[node] = current
                g_score[node] = tentative
                heuristic = math.hypot(nx - gx, ny - gy)
                heapq.heappush(open_heap, (tentative + heuristic, node))
    return sample_line(start, goal)


def build_grid_path(risk: np.ndarray, points: Sequence[Tuple[float, float]]) -> List[Tuple[int, int]]:
    full_path: List[Tuple[int, int]] = []
    for a, b in zip(points[:-1], points[1:]):
        seg = astar_path(risk, a, b)
        if full_path and seg and full_path[-1] == seg[0]:
            full_path.extend(seg[1:])
        else:
            full_path.extend(seg)
    return full_path


def evaluate_sequence_on_grid_path(
    sequence: Sequence[Target],
    risk: np.ndarray,
    start: Tuple[float, float],
    speed: float,
    fuel_consumption: float,
    max_time: float,
    service_time: float = 1.0,
    cell_size_km: float = DEFAULT_CELL_SIZE_KM,
) -> Tuple[RouteResult, List[Tuple[int, int]]]:
    """按A*避险网格路径重新评价路径，使结果指标与最终可执行轨迹一致。"""
    waypoints: List[Tuple[float, float]] = [start] + [(t.x, t.y) for t in sequence] + [start]
    full_path: List[Tuple[int, int]] = []
    segment_paths: List[List[Tuple[int, int]]] = []
    for a, b in zip(waypoints[:-1], waypoints[1:]):
        seg = astar_path(risk, a, b)
        segment_paths.append(seg)
        if full_path and seg and full_path[-1] == seg[0]:
            full_path.extend(seg[1:])
        else:
            full_path.extend(seg)

    total_time = 0.0
    total_distance = 0.0
    risk_cost = 0.0
    total_reward = 0.0
    expected_reward = 0.0
    survival_to_now = 1.0

    for idx, seg in enumerate(segment_paths):
        seg_distance = 0.0
        seg_risk_cost = 0.0
        for node_idx, (x, y) in enumerate(seg):
            if node_idx > 0:
                px, py = seg[node_idx - 1]
                seg_distance += math.hypot(x - px, y - py) * cell_size_km
            if node_idx > 0 or idx == 0:
                p = float(np.clip(risk[y, x], 0.0, 0.999))
                seg_risk_cost += -math.log(1.0 - p)
        total_distance += seg_distance
        total_time += seg_distance / speed
        risk_cost += seg_risk_cost
        survival_to_now *= math.exp(-seg_risk_cost)
        if idx < len(sequence):
            target = sequence[idx]
            total_time += service_time
            total_reward += target.weight
            expected_reward += target.weight * survival_to_now

    result = RouteResult(
        uav_id="",
        sequence=[t.target_id for t in sequence],
        time=total_time,
        reward=total_reward,
        expected_reward=expected_reward,
        risk_cost=risk_cost,
        survival=math.exp(-risk_cost),
        energy=fuel_consumption * total_distance,
        feasible=total_time <= max_time,
        path_points=waypoints,
    )
    return result, full_path


def write_csv(path: Path, rows: Sequence[Dict[str, object]], fieldnames: Sequence[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def save_targets_csv(targets: Sequence[Target], path: Path) -> None:
    write_csv(
        path,
        [
            {
                "target_id": t.target_id,
                "x": round(t.x, 4),
                "y": round(t.y, 4),
                "weight": round(t.weight, 6),
                "cell_count": t.cell_count,
                "local_risk": round(t.local_risk, 6),
                "service_time": t.service_time,
            }
            for t in targets
        ],
        ["target_id", "x", "y", "weight", "cell_count", "local_risk", "service_time"],
    )


def plot_risk_and_targets(risk: np.ndarray, targets: Sequence[Target], path: Path, title: str) -> None:
    setup_plot_style()
    fig, ax = plt.subplots(figsize=(8, 7))
    im = ax.imshow(risk, cmap="magma", origin="upper")
    xs = [t.x for t in targets]
    ys = [t.y for t in targets]
    sizes = [max(12, min(120, t.weight * 6)) for t in targets]
    ax.scatter(xs, ys, s=sizes, c="#3DB7E9", edgecolors="white", linewidths=0.4, alpha=0.85)
    ax.set_title(title)
    ax.set_xlabel("x / km")
    ax.set_ylabel("y / km")
    fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04, label="综合风险概率")
    fig.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def plot_all_targets_and_candidates(
    risk: np.ndarray,
    all_targets: Sequence[Target],
    candidate_targets: Sequence[Target],
    route_targets: Sequence[Target],
    path: Path,
    title: str,
) -> None:
    setup_plot_style()
    fig, ax = plt.subplots(figsize=(8, 7))
    ax.imshow(risk, cmap="Greys", origin="upper", alpha=0.55)
    ax.scatter(
        [t.x for t in all_targets],
        [t.y for t in all_targets],
        s=12,
        c="#B0B0B0",
        alpha=0.45,
        label="全部目标簇",
    )
    ax.scatter(
        [t.x for t in candidate_targets],
        [t.y for t in candidate_targets],
        s=32,
        c="#1F77B4",
        alpha=0.85,
        edgecolors="white",
        linewidths=0.3,
        label="可达/候选目标",
    )
    ax.scatter(
        [t.x for t in route_targets],
        [t.y for t in route_targets],
        s=70,
        c="#D62728",
        alpha=0.95,
        edgecolors="white",
        linewidths=0.5,
        label="实际访问目标",
    )
    ax.scatter([0], [0], marker="*", s=180, c="#2CA02C", edgecolors="white", linewidths=0.5, label="起降点")
    ax.set_title(title)
    ax.set_xlabel("x / 网格")
    ax.set_ylabel("y / 网格")
    ax.legend(loc="upper right", frameon=True)
    fig.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def plot_route(
    risk: np.ndarray,
    targets: Sequence[Target],
    route_points: Sequence[Tuple[float, float]],
    path: Path,
    title: str,
    zoom: bool = False,
    padding: float = 18.0,
) -> None:
    setup_plot_style()
    fig, ax = plt.subplots(figsize=(8, 7))
    ax.imshow(risk, cmap="Greys", origin="upper", alpha=0.65)
    ax.scatter([t.x for t in targets], [t.y for t in targets], s=20, c="#999999", alpha=0.35, label="候选目标")
    if route_points:
        xs = [p[0] for p in route_points]
        ys = [p[1] for p in route_points]
        ax.plot(xs, ys, color="#D62728", linewidth=2.2, label="规划路径")
        ax.scatter(xs[1:-1], ys[1:-1], s=45, color="#1F77B4", edgecolors="white", zorder=5, label="完成目标")
        ax.scatter([xs[0]], [ys[0]], marker="*", s=180, color="#2CA02C", edgecolors="white", zorder=6, label="起降点")
    ax.set_title(title)
    ax.set_xlabel("x / km")
    ax.set_ylabel("y / km")
    if zoom and route_points:
        xs = [p[0] for p in route_points]
        ys = [p[1] for p in route_points]
        xmin = max(0, min(xs) - padding)
        xmax = min(WIDTH - 1, max(xs) + padding)
        ymin = max(0, min(ys) - padding)
        ymax = min(HEIGHT - 1, max(ys) + padding)
        ax.set_xlim(xmin, xmax)
        ax.set_ylim(ymax, ymin)
    ax.legend(loc="upper right", frameon=True)
    fig.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def plot_grid_route(
    risk: np.ndarray,
    targets: Sequence[Target],
    route_targets: Sequence[Target],
    grid_path: Sequence[Tuple[int, int]],
    path: Path,
    title: str,
    zoom: bool = False,
    padding: float = 18.0,
) -> None:
    setup_plot_style()
    fig, ax = plt.subplots(figsize=(8, 7))
    ax.imshow(risk, cmap="Greys", origin="upper", alpha=0.65)
    ax.scatter([t.x for t in targets], [t.y for t in targets], s=18, c="#A8A8A8", alpha=0.32, label="全部目标簇")
    if grid_path:
        xs = [p[0] for p in grid_path]
        ys = [p[1] for p in grid_path]
        ax.plot(xs, ys, color="#D62728", linewidth=2.0, label="A*避险路径")
    if route_targets:
        ax.scatter(
            [t.x for t in route_targets],
            [t.y for t in route_targets],
            s=62,
            color="#1F77B4",
            edgecolors="white",
            linewidths=0.5,
            zorder=5,
            label="完成目标",
        )
    ax.scatter([0], [0], marker="*", s=180, color="#2CA02C", edgecolors="white", linewidths=0.5, zorder=6, label="起降点")
    ax.set_title(title)
    ax.set_xlabel("x / 网格")
    ax.set_ylabel("y / 网格")
    if zoom and grid_path:
        xs = [p[0] for p in grid_path]
        ys = [p[1] for p in grid_path]
        ax.set_xlim(max(0, min(xs) - padding), min(WIDTH - 1, max(xs) + padding))
        ax.set_ylim(min(HEIGHT - 1, max(ys) + padding), max(0, min(ys) - padding))
    ax.legend(loc="upper right", frameon=True)
    fig.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def plot_global_reachable_range(
    risk: np.ndarray,
    all_targets: Sequence[Target],
    route_targets: Sequence[Target],
    uav_id: str,
    start: Tuple[float, float],
    max_time: float,
    speed: float,
    cell_size_km: float,
    path: Path,
) -> None:
    setup_plot_style()
    fig, ax = plt.subplots(figsize=(8, 7))
    ax.imshow(risk, cmap="Greys", origin="upper", alpha=0.55)
    ax.scatter([t.x for t in all_targets], [t.y for t in all_targets], s=12, c="#B0B0B0", alpha=0.42, label="全部目标簇")
    ax.scatter(
        [t.x for t in route_targets],
        [t.y for t in route_targets],
        s=70,
        c="#D62728",
        edgecolors="white",
        linewidths=0.5,
        label="实际访问目标",
        zorder=4,
    )
    radius_grid = (max_time * speed / 2.0) / max(cell_size_km, 1e-9)
    circle = plt.Circle(start, radius_grid, fill=False, color="#1F77B4", linewidth=2.2, linestyle="--", label="理论往返可达边界")
    ax.add_patch(circle)
    ax.scatter([start[0]], [start[1]], marker="*", s=180, c="#2CA02C", edgecolors="white", linewidths=0.5, label="起降点")
    ax.set_xlim(0, WIDTH - 1)
    ax.set_ylim(HEIGHT - 1, 0)
    ax.set_title(f"问题1 全局目标簇与 {uav_id} 理论可达范围")
    ax.set_xlabel("x / 网格")
    ax.set_ylabel("y / 网格")
    ax.legend(loc="upper right", frameon=True)
    fig.tight_layout()
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def plot_problem1_diagnostics(targets: Sequence[Target], scale_rows: Sequence[Dict[str, object]]) -> None:
    setup_plot_style()
    weights = np.array([t.weight for t in targets], dtype=float)
    risks = np.array([t.local_risk for t in targets], dtype=float)

    fig, ax = plt.subplots(figsize=(7, 5))
    ax.hist(weights, bins=30, color="#4C78A8", edgecolor="white", alpha=0.9)
    ax.axvline(float(np.median(weights)), color="#E45756", linestyle="--", linewidth=2, label="中位数")
    ax.set_xlabel("目标簇权重")
    ax.set_ylabel("数量")
    ax.set_title("问题1 目标簇权重分布")
    ax.grid(axis="y", linestyle="--", alpha=0.4)
    ax.legend(frameon=True)
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "问题1_目标权重直方图.png", dpi=300, bbox_inches="tight")
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(7, 5))
    ax.hist(risks, bins=30, color="#72B7B2", edgecolor="white", alpha=0.9)
    ax.axvline(float(np.mean(risks)), color="#E45756", linestyle="--", linewidth=2, label="均值")
    ax.set_xlabel("目标簇局部风险")
    ax.set_ylabel("数量")
    ax.set_title("问题1 目标簇局部风险分布")
    ax.grid(axis="y", linestyle="--", alpha=0.4)
    ax.legend(frameon=True)
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "问题1_目标风险直方图.png", dpi=300, bbox_inches="tight")
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(7, 5))
    sizes = np.clip(weights * 2.0, 18, 160)
    sc = ax.scatter(risks, weights, s=sizes, c=[t.y for t in targets], cmap="viridis", alpha=0.72, edgecolors="white", linewidths=0.3)
    ax.set_xlabel("目标簇局部风险")
    ax.set_ylabel("目标簇权重")
    ax.set_title("问题1 权重-风险散点诊断")
    ax.grid(True, linestyle="--", alpha=0.35)
    fig.colorbar(sc, ax=ax, label="y 坐标")
    fig.tight_layout()
    fig.savefig(FIGURES_DIR / "问题1_权重风险散点图.png", dpi=300, bbox_inches="tight")
    plt.close(fig)

    if scale_rows:
        xs = [float(row["cell_size_km"]) for row in scale_rows]
        candidate_counts = [float(row["candidate_count"]) for row in scale_rows]
        completed = [float(row["completed_targets"]) for row in scale_rows]
        rewards = [float(row["expected_reward"]) for row in scale_rows]
        fig, ax1 = plt.subplots(figsize=(7, 5))
        ax1.plot(xs, candidate_counts, marker="o", color="#4C78A8", label="候选目标数")
        ax1.plot(xs, completed, marker="s", color="#F58518", label="完成目标数")
        ax1.set_xlabel("网格尺度 / km")
        ax1.set_ylabel("目标数量")
        ax1.grid(True, linestyle="--", alpha=0.35)
        ax2 = ax1.twinx()
        ax2.plot(xs, rewards, marker="^", color="#E45756", label="期望收益")
        ax2.set_ylabel("期望收益")
        lines1, labels1 = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        ax1.legend(lines1 + lines2, labels1 + labels2, loc="best", frameon=True)
        ax1.set_title("问题1 网格尺度敏感性")
        fig.tight_layout()
        fig.savefig(FIGURES_DIR / "问题1_网格尺度敏感性.png", dpi=300, bbox_inches="tight")
        plt.close(fig)


def load_preprocessed(threshold: float = 0.1, amplitude: float = 0.65, sigma: float = 18.0) -> Tuple[np.ndarray, np.ndarray, List[Tuple[int, int]], List[Target], List[Uav]]:
    base_risk, air_defense, value = load_base_data()
    fire_points = find_fire_points(air_defense)
    risk = build_risk_field(base_risk, fire_points, amplitude=amplitude, sigma=sigma)
    targets = extract_targets(value, risk, threshold=threshold)
    uavs = load_uavs()
    return risk, value, fire_points, targets, uavs


def target_lookup(targets: Sequence[Target]) -> Dict[int, Target]:
    return {t.target_id: t for t in targets}
